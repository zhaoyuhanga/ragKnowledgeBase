# -*- coding: utf-8 -*-
"""
表格解析器

本模块实现Excel和CSV表格的解析。
"""

from typing import Any, Dict, List, Optional

from app.parsers.base import (
    BaseParser,
    DocumentElementModel,
    ElementType,
    TableStructure,
)


class TableParser(BaseParser):
    """
    表格解析器

    支持Excel（.xlsx, .xls）和CSV文件的解析。
    """

    def __init__(self):
        """初始化表格解析器"""
        super().__init__()
        self.supported_extensions = ["xlsx", "xls", "csv"]

    def parse(
        self,
        file_path: str,
        version_id: int,
        document_id: int
    ) -> List[DocumentElementModel]:
        """
        解析表格文件

        Args:
            file_path: 文件路径
            version_id: 版本ID
            document_id: 文档ID

        Returns:
            解析后的元素列表
        """
        import os
        ext = os.path.splitext(file_path)[1].lower().lstrip(".")

        if ext in ["xlsx", "xls"]:
            return self._parse_excel(file_path, version_id, document_id)
        elif ext == "csv":
            return self._parse_csv(file_path, version_id, document_id)
        else:
            return []

    def _parse_excel(
        self,
        file_path: str,
        version_id: int,
        document_id: int
    ) -> List[DocumentElementModel]:
        """
        解析Excel文件

        Args:
            file_path: 文件路径
            version_id: 版本ID
            document_id: 文档ID

        Returns:
            表格元素列表
        """
        elements = []

        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, data_only=True)

            for sheet_index, sheet_name in enumerate(wb.sheetnames):
                sheet = wb[sheet_name]

                # 提取表头（第一行）
                headers = []
                for cell in sheet[1]:
                    headers.append(str(cell.value) if cell.value else "")

                # 提取数据行
                rows = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_data = [str(cell) if cell else "" for cell in row]
                    if any(row_data):  # 跳过空行
                        rows.append(row_data)

                # 检测合并单元格
                merged_cells = self._detect_merged_cells(sheet)

                # 创建表格结构
                table_structure = TableStructure(
                    headers=[headers] if headers else [],
                    rows=rows,
                    merged_cells=merged_cells,
                    row_count=len(rows) + (1 if headers else 0),
                    col_count=len(headers) if headers else 0,
                    caption=sheet_name
                )

                # 创建表格元素
                element = DocumentElementModel(
                    element_id=self._generate_element_id(),
                    document_id=document_id,
                    version_id=version_id,
                    page_no=sheet_index + 1,
                    element_type=ElementType.TABLE,
                    content=table_structure.to_text(),
                    reading_order=sheet_index,
                    table_structure=table_structure,
                    confidence=0.95
                )
                elements.append(element)

            wb.close()

        except ImportError:
            # openpyxl未安装
            import logging
            logger = logging.getLogger("rag.parser")
            logger.error("openpyxl未安装，无法解析Excel文件")
        except Exception as e:
            import logging
            logger = logging.getLogger("rag.parser")
            logger.error(f"Excel解析失败: {file_path}, 错误: {str(e)}")

        return elements

    def _parse_csv(
        self,
        file_path: str,
        version_id: int,
        document_id: int
    ) -> List[DocumentElementModel]:
        """
        解析CSV文件

        Args:
            file_path: 文件路径
            version_id: 版本ID
            document_id: 文档ID

        Returns:
            表格元素列表
        """
        elements = []

        try:
            import csv

            with open(file_path, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)

                # 读取所有行
                all_rows = list(reader)

                if not all_rows:
                    return elements

                # 第一行作为表头
                headers = all_rows[0] if all_rows else []
                data_rows = all_rows[1:] if len(all_rows) > 1 else []

                # 创建表格结构
                table_structure = TableStructure(
                    headers=[headers] if headers else [],
                    rows=data_rows,
                    merged_cells=[],
                    row_count=len(data_rows) + (1 if headers else 0),
                    col_count=len(headers) if headers else 0,
                    caption="CSV数据"
                )

                # 创建表格元素
                element = DocumentElementModel(
                    element_id=self._generate_element_id(),
                    document_id=document_id,
                    version_id=version_id,
                    page_no=1,
                    element_type=ElementType.TABLE,
                    content=table_structure.to_text(),
                    reading_order=0,
                    table_structure=table_structure,
                    confidence=0.95
                )
                elements.append(element)

        except Exception as e:
            import logging
            logger = logging.getLogger("rag.parser")
            logger.error(f"CSV解析失败: {file_path}, 错误: {str(e)}")

        return elements

    def _detect_merged_cells(self, sheet) -> List[Dict[str, Any]]:
        """
        检测合并单元格

        Args:
            sheet: Excel工作表对象

        Returns:
            合并单元格列表
        """
        merged_cells = []

        try:
            # 获取合并单元格范围
            if hasattr(sheet, "merged_cells"):
                for merged_range in sheet.merged_cells.ranges:
                    # 解析合并范围
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    merged_cells.append({
                        "min_col": min_col,
                        "min_row": min_row,
                        "max_col": max_col,
                        "max_row": max_row
                    })
        except Exception:
            pass

        return merged_cells
